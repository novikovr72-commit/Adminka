#!/usr/bin/env python3
import argparse
import os
import posixpath
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
import psycopg2
from psycopg2.extras import execute_values


DEFAULT_INPUT = "docs/Сотрудники - Все.xlsx"
DEFAULT_OUTPUT = "docs/Сотрудники - Все.with-ids.xlsx"
DEFAULT_CLAIMS_INPUT = "docs/Исходник_Claims.xlsx"
DEFAULT_CLAIMS_OUTPUT = "docs/Исходник_Claims.with-responsible-id.xlsx"

XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
XR_NS = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
XR2_NS = "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"
XR3_NS = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"

ET.register_namespace("", XML_NS)
ET.register_namespace("r", REL_NS)
ET.register_namespace("mc", MC_NS)
ET.register_namespace("x14ac", X14AC_NS)
ET.register_namespace("xr", XR_NS)
ET.register_namespace("xr2", XR2_NS)
ET.register_namespace("xr3", XR3_NS)


@dataclass
class EmployeeRow:
    row_index: int
    full_name: Optional[str]
    email: Optional[str]
    personal_number: Optional[int]
    surname: Optional[str]
    first_name: Optional[str]
    middle_name: Optional[str]
    sap_id: Optional[str]


def normalize_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return re.sub(r"\s+", " ", text)


def normalize_email(value: object) -> Optional[str]:
    text = normalize_text(value)
    if text is None:
        return None
    return text.lower()


def parse_personal_number(value: object) -> Optional[int]:
    text = normalize_text(value)
    if text is None:
        return None

    # В Excel номер может приходить как "12345.0".
    text = text.replace(",", ".")
    if text.endswith(".0"):
        text = text[:-2]
    if not text.isdigit():
        return None

    try:
        return int(text)
    except ValueError:
        return None


def split_fio(full_name: Optional[str]) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    if full_name is None:
        return None, None, None
    parts = [part for part in full_name.split(" ") if part]
    if not parts:
        return None, None, None
    surname = parts[0] if len(parts) >= 1 else None
    first_name = parts[1] if len(parts) >= 2 else None
    middle_name = " ".join(parts[2:]) if len(parts) >= 3 else None
    return surname, first_name, middle_name


def find_column_indexes(header_row) -> Dict[str, int]:
    indexes: Dict[str, int] = {}
    for idx, cell in enumerate(header_row, start=1):
        title = normalize_text(cell.value)
        if title is None:
            continue
        indexes[title] = idx
    return indexes


def resolve_column(index: Dict[str, int], *variants: str) -> Optional[int]:
    for title in variants:
        found = index.get(title)
        if found is not None:
            return found
    return None


def load_rows(
    input_path: Path,
) -> Tuple[openpyxl.Workbook, openpyxl.worksheet.worksheet.Worksheet, List[EmployeeRow], int]:
    workbook = openpyxl.load_workbook(input_path)
    sheet = workbook.worksheets[0]
    header = next(sheet.iter_rows(min_row=1, max_row=1))
    index = find_column_indexes(header)

    id_col = resolve_column(index, "ID", "Id", "id")
    ext_id_col = resolve_column(index, "Внешний ид.", "Внешний Ид.", "Внешний ИД.")
    fio_col = resolve_column(index, "ФИО")
    email_col = resolve_column(index, "E-mail", "Email", "E mail")
    personal_col = resolve_column(index, "Табельный номер", "Табельный №")

    required_pairs = [
        ("ID", id_col),
        ("Внешний ид.", ext_id_col),
        ("ФИО", fio_col),
        ("E-mail", email_col),
        ("Табельный номер", personal_col),
    ]
    missing = [name for name, value in required_pairs if value is None]
    if missing:
        raise RuntimeError(f"В файле отсутствуют обязательные колонки: {', '.join(missing)}")

    rows: List[EmployeeRow] = []
    for row_idx in range(2, sheet.max_row + 1):
        full_name = normalize_text(sheet.cell(row=row_idx, column=fio_col).value)  # type: ignore[arg-type]
        email = normalize_email(sheet.cell(row=row_idx, column=email_col).value)  # type: ignore[arg-type]
        personal_number = parse_personal_number(sheet.cell(row=row_idx, column=personal_col).value)  # type: ignore[arg-type]
        sap_id = normalize_text(sheet.cell(row=row_idx, column=ext_id_col).value)  # type: ignore[arg-type]
        surname, first_name, middle_name = split_fio(full_name)

        if full_name is None and email is None and personal_number is None and sap_id is None:
            continue
        rows.append(
            EmployeeRow(
                row_index=row_idx,
                full_name=full_name,
                email=email,
                personal_number=personal_number,
                surname=surname,
                first_name=first_name,
                middle_name=middle_name,
                sap_id=sap_id,
            )
        )

    return workbook, sheet, rows, id_col  # type: ignore[return-value]


def load_employee_id_map_for_claims(source_path: Path) -> Dict[str, str]:
    workbook = openpyxl.load_workbook(source_path)
    sheet = workbook.worksheets[0]
    header = next(sheet.iter_rows(min_row=1, max_row=1))
    index = find_column_indexes(header)

    fio_col = resolve_column(index, "ФИО")
    id_col = resolve_column(index, "ID", "Id", "id")
    if fio_col is None or id_col is None:
        raise RuntimeError("В файле сотрудников отсутствуют колонки ФИО/ID для сопоставления responsible_id")

    result: Dict[str, str] = {}
    for row_idx in range(2, sheet.max_row + 1):
        fio = normalize_text(sheet.cell(row=row_idx, column=fio_col).value)
        employee_id = normalize_text(sheet.cell(row=row_idx, column=id_col).value)
        if fio is None or employee_id is None:
            continue
        # Берем первое вхождение ФИО.
        if fio not in result:
            result[fio] = employee_id
    return result


def col_letters_to_index(letters: str) -> int:
    value = 0
    for char in letters:
        value = value * 26 + (ord(char) - ord("A") + 1)
    return value


def index_to_col_letters(index: int) -> str:
    value = index
    out = ""
    while value > 0:
        value, rem = divmod(value - 1, 26)
        out = chr(ord("A") + rem) + out
    return out


def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
    match = re.match(r"^([A-Z]+)(\d+)$", cell_ref)
    if not match:
        raise RuntimeError(f"Некорректная ссылка на ячейку: {cell_ref}")
    col = col_letters_to_index(match.group(1))
    row = int(match.group(2))
    return col, row


def parse_range_ref(range_ref: str) -> Tuple[int, int]:
    parts = range_ref.split(":")
    end_ref = parts[-1]
    col, row = parse_cell_ref(end_ref)
    return col, row


def first_sheet_xml_path(zip_file: ZipFile) -> str:
    workbook_root = ET.fromstring(zip_file.read("xl/workbook.xml"))
    sheets_node = workbook_root.find(f"{{{XML_NS}}}sheets")
    if sheets_node is None:
        raise RuntimeError("В workbook.xml отсутствует секция sheets")
    first_sheet = sheets_node.find(f"{{{XML_NS}}}sheet")
    if first_sheet is None:
        raise RuntimeError("В workbook.xml отсутствуют листы")
    rel_id = first_sheet.attrib.get(f"{{{REL_NS}}}id")
    if not rel_id:
        raise RuntimeError("У первого листа отсутствует relationship id")

    rels_root = ET.fromstring(zip_file.read("xl/_rels/workbook.xml.rels"))
    for rel in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
        if rel.attrib.get("Id") != rel_id:
            continue
        target = rel.attrib.get("Target")
        if not target:
            break
        normalized = target.lstrip("/")
        if not normalized.startswith("xl/"):
            normalized = "xl/" + normalized
        return normalized
    raise RuntimeError("Не удалось определить XML первого листа claims")


def first_sheet_table_xml_path(zip_file: ZipFile, sheet_path: str) -> Optional[str]:
    rels_path = sheet_path.replace("worksheets/", "worksheets/_rels/") + ".rels"
    if rels_path not in zip_file.namelist():
        return None
    rels_root = ET.fromstring(zip_file.read(rels_path))
    for rel in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship"):
        rel_type = rel.attrib.get("Type", "")
        if not rel_type.endswith("/table"):
            continue
        target = rel.attrib.get("Target")
        if not target:
            continue
        table_path = posixpath.normpath(posixpath.join("xl/worksheets", target))
        if table_path in zip_file.namelist():
            return table_path
    return None


def parse_shared_strings(zip_file: ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zip_file.namelist():
        return []
    root = ET.fromstring(zip_file.read("xl/sharedStrings.xml"))
    strings: List[str] = []
    for si in root.findall(f"{{{XML_NS}}}si"):
        text_parts = []
        direct_t = si.find(f"{{{XML_NS}}}t")
        if direct_t is not None and direct_t.text is not None:
            text_parts.append(direct_t.text)
        for run_t in si.findall(f".//{{{XML_NS}}}r/{{{XML_NS}}}t"):
            if run_t.text is not None:
                text_parts.append(run_t.text)
        strings.append("".join(text_parts))
    return strings


def cell_text_from_xml(cell: ET.Element, shared_strings: List[str]) -> Optional[str]:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        text_parts = []
        direct_t = cell.find(f"{{{XML_NS}}}is/{{{XML_NS}}}t")
        if direct_t is not None and direct_t.text is not None:
            text_parts.append(direct_t.text)
        for run_t in cell.findall(f".//{{{XML_NS}}}is/{{{XML_NS}}}r/{{{XML_NS}}}t"):
            if run_t.text is not None:
                text_parts.append(run_t.text)
        return "".join(text_parts) if text_parts else None

    value_node = cell.find(f"{{{XML_NS}}}v")
    if value_node is None or value_node.text is None:
        return None
    if cell_type == "s":
        try:
            idx = int(value_node.text)
            if 0 <= idx < len(shared_strings):
                return shared_strings[idx]
        except ValueError:
            return None
        return None
    return value_node.text


def set_inline_cell_value(row_node: ET.Element, cell_ref: str, value: str) -> None:
    desired_col, _ = parse_cell_ref(cell_ref)
    existing_cells = row_node.findall(f"{{{XML_NS}}}c")
    target_cell = None
    for cell in existing_cells:
        if cell.attrib.get("r") == cell_ref:
            target_cell = cell
            break

    if target_cell is None:
        target_cell = ET.Element(f"{{{XML_NS}}}c", {"r": cell_ref})
        insert_at = len(existing_cells)
        for index, cell in enumerate(existing_cells):
            cell_col, _ = parse_cell_ref(cell.attrib.get("r", "A1"))
            if cell_col > desired_col:
                insert_at = index
                break
        row_node.insert(insert_at, target_cell)

    target_cell.attrib["t"] = "inlineStr"
    for child in list(target_cell):
        target_cell.remove(child)
    is_node = ET.SubElement(target_cell, f"{{{XML_NS}}}is")
    t_node = ET.SubElement(is_node, f"{{{XML_NS}}}t")
    t_node.text = value

    # Поддерживаем корректный spans (min:max) для строки после добавления новой ячейки.
    cell_cols: List[int] = []
    for cell in row_node.findall(f"{{{XML_NS}}}c"):
        ref = cell.attrib.get("r")
        if not ref:
            continue
        col_idx, _ = parse_cell_ref(ref)
        cell_cols.append(col_idx)
    if cell_cols:
        row_node.attrib["spans"] = f"{min(cell_cols)}:{max(cell_cols)}"


def enrich_claims_responsible_id(
    claims_input_path: Path,
    claims_output_path: Path,
    fio_to_id: Dict[str, str],
    fallback_employee_id: str,
) -> Tuple[int, int]:
    with ZipFile(claims_input_path, "r") as source_zip:
        sheet_path = first_sheet_xml_path(source_zip)
        table_path = first_sheet_table_xml_path(source_zip, sheet_path)
        shared_strings = parse_shared_strings(source_zip)
        sheet_root = ET.fromstring(source_zip.read(sheet_path))
        # Убираем ссылки на префиксы, которые ElementTree может не сериализовать.
        sheet_root.set(f"{{{MC_NS}}}Ignorable", "x14ac xr")

        sheet_data = sheet_root.find(f"{{{XML_NS}}}sheetData")
        if sheet_data is None:
            raise RuntimeError("В claims-файле отсутствует секция sheetData")

        row_nodes = sheet_data.findall(f"{{{XML_NS}}}row")
        if not row_nodes:
            raise RuntimeError("В claims-файле отсутствуют строки")

        header_row = None
        rows_by_index: Dict[int, ET.Element] = {}
        max_row_index = 1
        for row in row_nodes:
            row_index = int(row.attrib.get("r", "0"))
            rows_by_index[row_index] = row
            if row_index == 1:
                header_row = row
            if row_index > max_row_index:
                max_row_index = row_index
        if header_row is None:
            raise RuntimeError("В claims-файле отсутствует строка заголовков")

        header_cols: Dict[str, int] = {}
        max_col_index = 1
        for cell in header_row.findall(f"{{{XML_NS}}}c"):
            cell_ref = cell.attrib.get("r")
            if not cell_ref:
                continue
            col_idx, _ = parse_cell_ref(cell_ref)
            if col_idx > max_col_index:
                max_col_index = col_idx
            title = normalize_text(cell_text_from_xml(cell, shared_strings))
            if title is not None:
                header_cols[title] = col_idx

        responsible_fio_col = header_cols.get("responsible_full_name")
        if responsible_fio_col is None:
            raise RuntimeError("В файле Исходник_Claims.xlsx отсутствует колонка responsible_full_name")

        responsible_id_col = header_cols.get("responsible_id")
        if responsible_id_col is None:
            responsible_id_col = max_col_index + 1
            set_inline_cell_value(header_row, f"{index_to_col_letters(responsible_id_col)}1", "responsible_id")
            max_col_index = responsible_id_col

        matched = 0
        fallback_filled = 0
        responsible_fio_col_letters = index_to_col_letters(responsible_fio_col)
        responsible_id_col_letters = index_to_col_letters(responsible_id_col)

        for row_idx in range(2, max_row_index + 1):
            row_node = rows_by_index.get(row_idx)
            if row_node is None:
                continue
            source_ref = f"{responsible_fio_col_letters}{row_idx}"
            source_value = None
            for cell in row_node.findall(f"{{{XML_NS}}}c"):
                if cell.attrib.get("r") == source_ref:
                    source_value = normalize_text(cell_text_from_xml(cell, shared_strings))
                    break

            employee_id = fio_to_id.get(source_value) if source_value is not None else None
            if employee_id is None:
                employee_id = fallback_employee_id
                fallback_filled += 1
            else:
                matched += 1

            target_ref = f"{responsible_id_col_letters}{row_idx}"
            set_inline_cell_value(row_node, target_ref, employee_id)

        last_col_index = max(max_col_index, responsible_id_col)
        last_col_letters = index_to_col_letters(last_col_index)
        dimension_node = sheet_root.find(f"{{{XML_NS}}}dimension")
        new_dimension_ref = f"A1:{last_col_letters}{max_row_index}"
        if dimension_node is None:
            dimension_node = ET.Element(f"{{{XML_NS}}}dimension")
            sheet_root.insert(0, dimension_node)
        dimension_node.attrib["ref"] = new_dimension_ref

        # На этом листе фильтрация идет через объект таблицы table1.xml.
        # Не добавляем worksheet.autoFilter, чтобы не ломать metadata таблицы.
        auto_filter_node = sheet_root.find(f"{{{XML_NS}}}autoFilter")
        if auto_filter_node is not None:
            sheet_root.remove(auto_filter_node)

        updated_sheet_xml = ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True)
        updated_table_xml: Optional[bytes] = None
        if table_path is not None:
            table_root = ET.fromstring(source_zip.read(table_path))
            table_original_ref = table_root.attrib.get("ref", new_dimension_ref)
            table_end_col, _ = parse_range_ref(table_original_ref)
            table_end_col = max(table_end_col, responsible_id_col)
            table_ref = f"A1:{index_to_col_letters(table_end_col)}{max_row_index}"
            table_root.attrib["ref"] = table_ref

            table_auto_filter = table_root.find(f"{{{XML_NS}}}autoFilter")
            if table_auto_filter is None:
                table_auto_filter = ET.Element(f"{{{XML_NS}}}autoFilter")
                table_root.insert(0, table_auto_filter)
            table_auto_filter.attrib["ref"] = table_ref

            table_columns = table_root.find(f"{{{XML_NS}}}tableColumns")
            if table_columns is not None:
                existing_columns = table_columns.findall(f"{{{XML_NS}}}tableColumn")
                has_responsible_id = any((col.attrib.get("name") or "") == "responsible_id" for col in existing_columns)
                if not has_responsible_id:
                    max_id = 0
                    for col in existing_columns:
                        raw_id = col.attrib.get("id")
                        if raw_id and raw_id.isdigit():
                            max_id = max(max_id, int(raw_id))
                    ET.SubElement(
                        table_columns,
                        f"{{{XML_NS}}}tableColumn",
                        {"id": str(max_id + 1), "name": "responsible_id"},
                    )
                table_columns.attrib["count"] = str(len(table_columns.findall(f"{{{XML_NS}}}tableColumn")))

            updated_table_xml = ET.tostring(table_root, encoding="utf-8", xml_declaration=True)

        claims_output_path.parent.mkdir(parents=True, exist_ok=True)
        with ZipFile(claims_output_path, "w", compression=ZIP_DEFLATED) as output_zip:
            for info in source_zip.infolist():
                data = source_zip.read(info.filename)
                if info.filename == sheet_path:
                    data = updated_sheet_xml
                elif table_path is not None and updated_table_xml is not None and info.filename == table_path:
                    data = updated_table_xml
                output_zip.writestr(info, data)

        return matched, fallback_filled


def open_connection(args: argparse.Namespace):
    return psycopg2.connect(
        host=args.db_host,
        port=args.db_port,
        dbname=args.db_name,
        user=args.db_user,
        password=args.db_password,
    )


def fetch_existing_by_emails(conn, emails: List[str]) -> Dict[str, str]:
    if not emails:
        return {}
    with conn.cursor() as cur:
        cur.execute(
            """
            select id::text, lower(email) as email
            from party.employee
            where email is not null
              and lower(email) = any(%s)
            """,
            (emails,),
        )
        return {row[1]: row[0] for row in cur.fetchall()}


def set_deleted_false(conn, emails: List[str]) -> None:
    if not emails:
        return
    with conn.cursor() as cur:
        cur.execute(
            """
            update party.employee
               set deleted = false,
                   updated_at = now()
             where email is not null
               and lower(email) = any(%s)
            """,
            (emails,),
        )


def insert_new_rows(conn, rows: List[EmployeeRow]) -> None:
    if not rows:
        return
    values = [
        (
            row.full_name,
            row.email,
            row.surname,
            row.first_name,
            row.middle_name,
            row.personal_number,
            row.sap_id,
        )
        for row in rows
    ]
    with conn.cursor() as cur:
        execute_values(
            cur,
            """
            insert into party.employee (
                full_name,
                email,
                surname,
                first_name,
                middle_name,
                personal_number,
                sap_id,
                deleted,
                created_at,
                updated_at
            ) values %s
            """,
            values,
            template="(%s, %s, %s, %s, %s, %s, %s, false, now(), now())",
        )


def update_sap_id_for_existing(conn, rows: List[EmployeeRow], existing_emails: set) -> None:
    updates: Dict[str, str] = {}
    for row in rows:
        if not row.email or not row.sap_id:
            continue
        if row.email not in existing_emails:
            continue
        if row.email in updates:
            continue
        updates[row.email] = row.sap_id

    if not updates:
        return

    with conn.cursor() as cur:
        for email, sap_id in updates.items():
            cur.execute(
                """
                update party.employee
                   set sap_id = %s,
                       deleted = false,
                       updated_at = now()
                 where email is not null
                   and lower(email) = %s
                """,
                (sap_id, email),
            )


def get_required_fallback_id(conn, fallback_email: str) -> str:
    with conn.cursor() as cur:
        cur.execute(
            """
            select id::text
            from party.employee
            where lower(email) = %s
            order by updated_at desc nulls last
            limit 1
            """,
            (fallback_email.lower(),),
        )
        row = cur.fetchone()
        if not row or not row[0]:
            raise RuntimeError(f"Не найдена запись в party.employee для fallback email: {fallback_email}")
        return row[0]


def process_rows(conn, rows: List[EmployeeRow]) -> Dict[str, str]:
    emails = sorted({row.email for row in rows if row.email})
    existing = fetch_existing_by_emails(conn, emails)

    existing_emails = sorted(existing.keys())
    if existing_emails:
        set_deleted_false(conn, existing_emails)
        update_sap_id_for_existing(conn, rows, set(existing_emails))

    # В исходном файле может быть несколько строк с одинаковым email.
    # Для вставки берём только первую, чтобы не нарушать unique(email).
    seen_new_emails = set()
    new_rows: List[EmployeeRow] = []
    for row in rows:
        if not row.email or row.email in existing:
            continue
        if row.email in seen_new_emails:
            continue
        seen_new_emails.add(row.email)
        new_rows.append(row)
    insert_new_rows(conn, new_rows)

    # После вставки запрашиваем id заново по всем email из исходного файла.
    return fetch_existing_by_emails(conn, emails)


def write_ids_and_save(
    workbook: openpyxl.Workbook,
    sheet,
    rows: List[EmployeeRow],
    id_col: int,
    email_to_id: Dict[str, str],
    fallback_employee_id: str,
    output_path: Path,
) -> None:
    for row in rows:
        employee_id = None
        if row.email:
            employee_id = email_to_id.get(row.email)
        if not employee_id and not row.email:
            employee_id = fallback_employee_id
        if employee_id:
            sheet.cell(row=row.row_index, column=id_col).value = employee_id

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Импорт сотрудников из Excel в party.employee и обратная выгрузка с id."
    )
    parser.add_argument("--input", default=DEFAULT_INPUT, help="Путь к исходному xlsx-файлу")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Путь к выходному xlsx-файлу")
    parser.add_argument(
        "--claims-input",
        default=DEFAULT_CLAIMS_INPUT,
        help="Путь к файлу Исходник_Claims.xlsx",
    )
    parser.add_argument(
        "--claims-output",
        default=DEFAULT_CLAIMS_OUTPUT,
        help="Путь к выходному файлу claims с колонкой responsible_id",
    )

    parser.add_argument("--db-host", default=os.getenv("DB_HOST", "localhost"))
    parser.add_argument("--db-port", type=int, default=int(os.getenv("DB_PORT", "5432")))
    parser.add_argument("--db-name", default=os.getenv("DB_NAME", "nlmk_test"))
    parser.add_argument("--db-user", default=os.getenv("DB_USER", "roman"))
    parser.add_argument("--db-password", default=os.getenv("DB_PASSWORD", ""))
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    input_path = Path(args.input)
    output_path = Path(args.output)
    if not input_path.exists():
        raise RuntimeError(f"Файл не найден: {input_path}")
    claims_input_path = Path(args.claims_input)
    if not claims_input_path.exists():
        raise RuntimeError(f"Файл не найден: {claims_input_path}")
    claims_output_path = Path(args.claims_output)

    workbook, sheet, rows, id_col = load_rows(input_path)
    if not rows:
        raise RuntimeError("В файле нет строк для обработки")

    with open_connection(args) as conn:
        conn.autocommit = False
        email_to_id = process_rows(conn, rows)
        fallback_employee_id = get_required_fallback_id(conn, "golovina_mn@nlmk.com")
        conn.commit()

    write_ids_and_save(workbook, sheet, rows, id_col, email_to_id, fallback_employee_id, output_path)
    fio_to_id = load_employee_id_map_for_claims(output_path)
    matched_claims, fallback_claims = enrich_claims_responsible_id(
        claims_input_path,
        claims_output_path,
        fio_to_id,
        fallback_employee_id,
    )

    print(f"Готово. Обработано строк: {len(rows)}")
    print(f"Результат сохранен: {output_path}")
    print(f"Claims обновлен. Совпадений по responsible_full_name: {matched_claims}")
    print(f"Claims обновлен. Заполнено fallback ID: {fallback_claims}")
    print(f"Файл claims сохранен: {claims_output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
